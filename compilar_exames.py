import pdfplumber
import pytesseract
from PIL import Image, ImageFilter, ImageOps
import pandas as pd
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import sys
from pathlib import Path
from difflib import SequenceMatcher
# To preprocess image
import cv2
import numpy as np

# === PREPROCESS IMAGE ===
def preprocess_image(image_path):
    # Load image
    img = cv2.imread(str(image_path))

    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Increase contrast using CLAHE (keep this - it was working)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    contrast = clahe.apply(gray)

    # Lighter blur to preserve small characters like slashes
    blurred = cv2.GaussianBlur(contrast, (3, 3), 0)

    # Use Otsu thresholding but be less aggressive
    _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # Lighter morphology to preserve date separators
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))  # Smaller kernel
    cleaned = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)

    # Skip median blur to preserve detail
    processed = cleaned

    # Resize large images
    height, width = processed.shape
    max_width = 2000
    if width > max_width:
        scale = max_width / width
        processed = cv2.resize(
            processed,
            (int(width * scale), int(height * scale)),
            interpolation=cv2.INTER_CUBIC  # Better for text
        )

    return processed

# === CONFIGURAÇÕES ===
if getattr(sys, 'frozen', False):
    base_path = Path(sys.executable).parent  # pasta do exe
else:
    base_path = Path(__file__).parent        # pasta do .py

pasta_arquivos = base_path / "exames"
saida_planilha = base_path / "resultados.xlsx"


# === MAPA DE EXAMES (com synonyms, group e título padronizado) ===
map_exames = {
    "25-HIDROXIVITAMINA D": {"synonyms": ["Hidroxivitamina D"], "group": "Vitaminas", "title": "Vit D"},
    "ANTIESTREPTOLISINAO": {"synonyms": ["Antiestreptolisina O"], "group": "Hematimétrico", "title": "AntEst", "qualitative": True},
    "HEMOSSEDIMENTACAO": {"synonyms": ["Hemossedimentação"], "group": "Hematimétrico", "title": "Hemos"},
    "HEMOGLOBINA GLICADA": {"synonyms": ["Hemoglobina glicada"], "group": "Hematimétrico", "title": "H1ac"},
    "HEMOGLOBINA": {"synonyms": ["Hemoglobina"], "group": "Hematimétrico", "title": "Hb"},
    "RETICULOCITOS": {"synonyms": ["Reticulócitos"], "group": "Hematimétrico", "title": "Ret"},
    "HEMATOCRITO": {"synonyms": ["Hematócrito"], "group": "Hematimétrico", "title": "Hmt"},
    "VCM": {"synonyms": ["VCM", "V.C.M.", "VGM"], "group": "Hematimétrico", "title": "VCM"},
    "HCM": {"synonyms": ["HCM"], "group": "Hematimétrico", "title": "HCM"},
    "CHCM": {"synonyms": ["CHCM"], "group": "Hematimétrico", "title": "CHCM"},
    "RDW": {"synonyms": ["RDW"], "group": "Hematimétrico", "title": "RDW"},    
    "PLAQUETAS (FONIO)": {"synonyms": ["Fônio"], "group": "Hematimétrico", "title": "Plq f"},
    "PLAQUETAS": {"synonyms": ["Plaquetas"], "group": "Hematimétrico", "title": "Plq"},    
    "VPM": {"synonyms": ["VPM"], "group": "Hematimétrico", "title": "VPM"},
    "LEUCOCITOS": {"synonyms": ["Leucócitos", "WBC"], "group": "Defesa", "title": "Leuc"},
    "ERITOBLASTOS": {"synonyms": ["Eritoblástos"], "group": "Hematimétrico", "title": "Erit"},
    "BLASTOS": {"synonyms": ["Blastos"], "group": "Defesa", "title": "Blas"},
    "BASTONETES": {"synonyms": ["Bastonetes"], "group": "Defesa", "title": "Basto"},
    "PROMIELOCITOS": {"synonyms": ["Promielócitos"], "group": "Defesa", "title": "Prom"},
    "MIELOCITOS": {"synonyms": ["Mielócitos"], "group": "Defesa", "title": "Miel"},
    "METAMIELOCITOS": {"synonyms": ["Metamielócitos"], "group": "Defesa", "title": "Metam"},    
    "SEGMENTADOS": {"synonyms": ["Segmentados", "Neutrófilos"], "group": "Defesa", "title": "Seg"},
    "LINFOCITOS": {"synonyms": ["Linfócitos", "Linfócitos típicos"], "group": "Defesa", "title": "Linf"},
    "LINFOCITOS ATIPICOS": {"synonyms": ["Linfócitos atípicos"], "group": "Defesa", "title": "L atí"},
    "MONOCITOS": {"synonyms": ["Monócitos"], "group": "Defesa", "title": "Mon"},
    "EOSINOFILOS": {"synonyms": ["Eosinófilos"], "group": "Defesa", "title": "Eos"},
    "BASOFILOS": {"synonyms": ["Basófilos"], "group": "Defesa", "title": "Bas"},
    "SODIO": {"synonyms": ["Sódio"], "group": "Eletrólitos", "title": "Na"},
    "POTASSIO": {"synonyms": ["Potássio"], "group": "Eletrólitos", "title": "K"},
    "MAGNESIO": {"synonyms": ["Magnésio"], "group": "Eletrólitos", "title": "Mg"},
    "CLORETOS": {"synonyms": ["Cloretos"], "group": "Eletrólitos", "title": "Cl"},
    "UREIA": {"synonyms": ["Ureia", "Uréia"], "group": "Função Renal", "title": "Ur"},
    "CREATININA": {"synonyms": ["Creatinina"], "group": "Função Renal", "title": "Cr"},
    "FILTRAÇÃO GLOMERULAR": {"synonyms": ["Filtração glomerular"], "group": "Função Renal", "title": "TFG"},
    "PROTEINA C REATIVA": {"synonyms": ["Proteína C Reativa"], "group": "Inflamação", "title": "PCR"},
    "BILIRRUBINA TOTAL": {"synonyms": ["Bilirrubina Total"], "group": "Função Hepática", "title": "BT"},
    "BILIRRUBINA DIRETA": {"synonyms": ["Bilirrubina Direta"], "group": "Função Hepática", "title": "BD"},
    "BILIRRUBINA INDIRETA": {"synonyms": ["Bilirrubina Indireta"], "group": "Função Hepática", "title": "BI"},
    "AST": {"synonyms": ["TRANSAMINASE OXALACETICA"], "group": "Função Hepática", "title": "AST"},
    "ALT": {"synonyms": ["TRANSAMINASE PIRUVICA"], "group": "Função Hepática", "title": "ALT"},
    "GAMA GT": {"synonyms": ["Gama GT"], "group": "Função Hepática", "title": "GGT"},
    "LDH": {"synonyms": ["LDH"], "group": "Função Hepática", "title": "LDH"},
    "HAPTOGLOBINA": {"synonyms": ["Haptoglobina"], "group": "Função Hepática", "title": "Hapt"},
    "PH": {"synonyms": ["pH"], "group": "Gasometria", "title": "pH"},
    "PO2": {"synonyms": ["pO2"], "group": "Gasometria", "title": "pO2"},
    "PCO2": {"synonyms": ["pCO2"], "group": "Gasometria", "title": "pCO2"},
    "HCO3": {"synonyms": ["HCO3", "Bicarbonato"], "group": "Gasometria", "title": "HCO3"},
    "CTCO2": {"synonyms": ["CTCO2"], "group": "Gasometria", "title": "CTCO2"},
    #"BE": {"synonyms": ["BE", "Excesso de Base"], "group": "Gasometria", "title": "BE"},
    "SO2": {"synonyms": ["SO2", "O2Sat"], "group": "Gasometria", "title": "SO2"},
    "LACTATO": {"synonyms": ["Lactato", "Ácido latico"], "group": "Gasometria", "title": "Lact"},
    "PROERITOBLASTOS": {"synonyms": ["Proeritoblastos"], "group": "Hematopoiese", "title": "Proeri"},
    "POLICROMATICOS": {"synonyms": ["Policromáticos"], "group": "Hematopoiese", "title": "Polic"},
    "ORTOCROMATICOS": {"synonyms": ["Ortocromáticos"], "group": "Hematopoiese", "title": "Ortoc"},
    "HAIRY CELLS": {"synonyms": ["Hairy Cells"], "group": "Hematopoiese", "title": "H Cel"},
    "VITAMINA B12": {"synonyms": ["Vitamina B12"], "group": "Gasometria", "title": "V B12"},
    "HOMOCISTEINA": {"synonyms": ["Homocisteína"], "group": "Gasometria", "title": "Homoc"},
    "ACIDO FOLICO": {"synonyms": ["Ácido fólico"], "group": "Gasometria", "title": "Ác Fól"},
    "TROPONINA": {"synonyms": ["Troponina", "Troponina I"], "group": "Miocárdio", "title": "Trop", "qualitative": True},
    "CPK": {"synonyms": ["CPK"], "group": "Miocárdio", "title": "CPK"},
    "CKMB": {"synonyms": ["CKMB"], "group": "Miocárdio", "title": "CKMB"},
    "ATIVIDADE DE PROTROMBINA": {"synonyms": ["Atividade"], "group": "Coagulograma", "title": "A Prot"},
    "TEMPO DE PROTROMBINA": {"synonyms": ["Tempo"], "group": "Coagulograma", "title": "T Prot"},    
    "RNI": {"synonyms": ["R.N.I"], "group": "Coagulograma", "title": "RNI"},
    "TEMPO DE TROMBOPLASTINA PARCIAL ATIVADO": {"synonyms": ["TEMPO DE TROMBOPLASTINA"], "group": "Coagulograma", "title": "TTPA"},
    "FERRO SERICO": {"synonyms": ["Ferro Sérico"], "group": "Ferro", "title": "Fe Sér"},
    #FATOR RH/DU
    #GRUPO SANGUINEO    
}

# Qualitative values for positive, negative, greater or smaller than
qualitative_values = {
    r"MENOR\s+QUE\s*0[,.]1": 0.05,
    r"NEGATIVO": 0.05,
    r"POSITIVO": 0.15,
    r"MAIOR\s+OU\s+IGUAL\s+A\s*0[,.]1": 0.15,
    r"INDETECT[ÁA]VEL": 0.0,
    r"DETECT[ÁA]VEL": 0.15
}

# === NORMALIZAÇÃO ===
def normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-Z0-9]", "", s.upper())

# === BUILD LOOKUP ===
exames_lookup = {}
for key, meta in map_exames.items():
    for s in [key] + meta["synonyms"]:        
        exames_lookup[normalize(s)] = meta
        #print(f"🔑 Lookup add: {s} -> {exames_lookup[normalize(s)]}")

# === GRUPOS AUTOMÁTICOS ===
grupos = {}
for meta in map_exames.values():
    grupos.setdefault(meta["group"], []).append(meta["title"])

# === FUNÇÃO PARA EXTRAÇÃO DE TEXTO === 
def extrair_texto(arquivo):
    is_ocr = False
    if arquivo.suffix.lower() == ".pdf":
        with pdfplumber.open(arquivo) as pdf:
            return "\n".join(page.extract_text() for page in pdf.pages if page.extract_text()), is_ocr

    elif arquivo.suffix.lower() in [".jpg", ".jpeg", ".png"]:
        # Simple approach - try different PSM modes without complex preprocessing
        processed_img = preprocess_image(arquivo)
        img_pil = Image.fromarray(processed_img)
        
        # Try different page segmentation modes
        psm_modes = [
            (6, 'Uniform block of text'),
            (4, 'Single column of text'),
            (3, 'Fully automatic page segmentation'),
            (8, 'Single word'),
            (13, 'Raw line')
        ]
        
        best_text = ""
        best_score = 0
        
        for psm, description in psm_modes:
            try:
                #config = f'--oem 3 --psm {psm} -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789.:/()- '
                config = f'--oem 3 --psm {psm}'
                texto = pytesseract.image_to_string(img_pil, lang="por", config=config)
                
                # Simple scoring based on medical terms
                score = 0
                medical_terms = ["paciente", "resultado", "exame", "laboratorio", "medico", "coleta"]
                for term in medical_terms:
                    if term in texto.lower():
                        score += 10
                
                if len(texto.strip()) > 50:
                    score += 5
                
                print(f"  PSM {psm} ({description}): score {score}")
                
                if score > best_score:
                    best_score = score
                    best_text = texto
                    
            except Exception as e:
                print(f"  PSM {psm} failed: {e}")
                continue
        
        # If nothing good found, try default
        if best_score == 0:
            best_text = pytesseract.image_to_string(img_pil, lang="por")
        
        print(f"=== BEST OCR TEXT ===")
        print(best_text)
        print(f"=== END OCR TEXT ===")
        
        return best_text, True
    
    return "", False

# === EXTRAÇÃO DE NOVOS DADOS ===
dados_novos = {}

for arquivo in pasta_arquivos.iterdir():
    if not arquivo.suffix.lower() in [".pdf", ".jpg", ".jpeg", ".png"]:
        continue    

    texto, is_ocr = extrair_texto(arquivo)
    print(f"\n Texto: {texto}")
    if not texto:
        continue

    linhas = re.split(r'\r\n|\r|\n', texto)
    linhas = [l.strip() for l in linhas if l.strip()]
    
    # Patient detection
    paciente = "Desconhecido"
    for i, linha in enumerate(linhas):
        linha_upper = linha.upper()
        
        # Better patient detection        
        if "PACIENTE" in linha_upper:
            paciente = re.split(r'(?i)paciente[:]?\s*', linha, 1)[-1].strip()
            # Remove metadata
            #paciente = re.split(r'(?i)\s+(?:DT|NASC|SEXO|RG|CPF|DATA|ATENDIMENTO|PEDIDO|CONVENIO|CONVÊNIO|CASU|PRONTU[AÁ]RIO)', paciente)[0]
            paciente = re.split(
                r'(?i)(?:DT|NASC|SEXO|RG|CPF|DATA|ATENDIMENTO|PEDIDO|CONVENIO|CONVÊNIO|CASU|PRONTU[AÁ]RIO)',
                paciente
            )[0]

            print(f"\n👤 Paciente limpo 1: {paciente}")
        
        if not paciente and i + 1 < len(linhas):
            paciente = linhas[i + 1].strip()
            paciente = re.split(r'(?i)\s+(?:DT|NASC|SEXO|RG|CPF|DATA|CASU)', paciente)[0]
            paciente = re.sub(r'(\w)\d+', r'\1', paciente)
            paciente = re.sub(r'[\d\-\.\(\)]+', ' ', paciente)
            paciente = re.sub(r'\s+', ' ', paciente).strip().upper()
        
        break

        # Clean up the patient name        
        paciente = re.sub(r'[^A-ZÀ-Ü\s]', ' ', paciente)
        paciente = re.sub(r'\s+', ' ', paciente).strip()
                
        print(f"\n👤 Paciente limpo 2: {paciente}")
        if paciente and len(paciente) > 2:
            break

    print(f"\n👤 Paciente detectado: {paciente}")

    exames_processados = set()  # to not overwrite added value with exam historic data

    # Try to find the exam date first (more specific patterns)
    date_patterns = [
        r"DATA\s*DO\s*ATENDIMENTO[:\s]+(\d{2}/\d{2}/\d{4})",
        r"DATA\s*DO\s*PEDIDO[:\s]+(\d{2}/\d{2}/\d{4})", 
        r"DATA\s*DA\s*COLETA[:\s]+(\d{2}/\d{2}/\d{4})",
        r"DATA\s*DE\s*COLETA[:\s]+(\d{2}/\d{2}/\d{4})",
        r"DATA\s*DO\s*EXAME[:\s]+(\d{2}/\d{2}/\d{4})"
    ]

    data = "00/00/0000"
    for pattern in date_patterns:
        m_data = re.search(pattern, texto, re.IGNORECASE)
        if m_data:
            data = m_data.group(1)
            print(f"✅ Data do exame encontrada: {data}")
            break

    # If no exam date found, look for any date but avoid birth date
    if data == "00/00/0000":
        all_dates = re.findall(r"(\d{2}/\d{2}/\d{4})", texto)
        for found_date in all_dates:
            # Skip dates that look like birth years (before 2000)
            year = int(found_date.split("/")[2])
            if year >= 2000:  # Assume exam dates are from 2000 onwards
                data = found_date
                print(f"⚠️  Data aproximada (não birth date): {data}")
                break

    for i, linha in enumerate(linhas):
        linha = re.sub(r'(?<=\d)\s+(?=\d)', '', linha)
        #print(f"\n👤 Linha: {linha}")
        linha_norm = normalize(linha)
        #print(f"\n👤 Linha_norm: {linha_norm}")

        # DEBUG: Check specifically for TROPONINA
        if "TROPONINA" in linha_norm:
            print(f"  🔍 TROPONINA DETECTED! Line: {linha}")
            print(f"  🔍 Normalized: {linha_norm}")
            print(f"  🔍 In lookup: {linha_norm in exames_lookup}")
            if linha_norm in exames_lookup:
                print(f"  🔍 Lookup result: {exames_lookup[linha_norm]}")
        
        # Skip non-lab lines (add this function)
        def is_lab_result_line(line):
            line = line.strip()
            if len(line) < 3:
                return False
            if any(marker in line.upper() for marker in ['PACIENTE:', 'RG:', 'DATA:', 'MATERIAL:', 'METODO:', 
                'LIBERAÇÃO:', 'REFERÊNCIA:', 'Coleta:', 'CEP:', 'DATA', 'IDENTIDADE', 'PRONTUARIO', 'ANS-', 'ATENDIMENTO', 'NASC', 'Hash:', 'Nota', 'doenca', 'CANCER', 'sindrome', 'risco']):
                return False
            
            # Accept lines that look like exam names (even without numbers)
            if re.search(r'[A-ZÀ-ÿ]{3,}', line, re.IGNORECASE):
                return True
            
            # Also accept lines with numbers (your existing logic)
            if re.search(r'[A-ZÀ-ÿ]{2,}.*\d+[,.]?\d*', line, re.IGNORECASE):
                return True
            if re.search(r'\d+[,.]?\d*\s*[A-ZÀ-ÿ]{2,}', line, re.IGNORECASE):
                return True

            # NEW: Check if this looks like an exam name with spaces (like "S O D I O")
            # Remove spaces and check if it matches known exam patterns
            no_space_line = re.sub(r'\s+', '', line)
            if re.search(r'^[A-ZÀ-ÿ]{3,}$', no_space_line, re.IGNORECASE):
                return True
            
            return False
        
        if not is_lab_result_line(linha):
            continue

        # ---------- Identificar melhor exame para esta linha ----------
        best_match = None
        best_ratio = 0
        best_exam_name = ""
        column_title = None
        valor = None
        linha_usada = False  # impede múltiplos exames por linha

        # Lista de hemograma diferencial
        hemograma_abs = [
            "Basofilos", "Eosinofilos", "Mielocitos", "Metamielocitos", "Promielocitos",
            "Bastonetes", "Segmentados", "Linfocitos Tipicos", "Plaquetas", "Bastonetes",
            "Linfocitos", "Monocitos", "Linfocitos Reativos", "Blastos", "Leucocitos", "ATIVIDADE DE PROTROMBINA", "Tempo DE PROTROMBINA", "RNI"
        ]
        hemograma_abs_norm = [normalize(x) for x in hemograma_abs]

        exam_part = re.sub(r'\d', '', linha[:20])
        exam_part = normalize(exam_part)

        for chave_norm, meta in exames_lookup.items():
            if not is_ocr:
                # Para PDF/texto limpo → aceita nome ou sinônimo que apareça na linha
                nomes_possiveis = [normalize(s) for s in [chave_norm] + meta["synonyms"]]
                if any(n in linha_norm for n in nomes_possiveis):
                    if best_match is None or len(chave_norm) > len(best_match[0]):
                        best_match = (chave_norm, meta)
                        best_exam_name = exam_part
                        break  # pega o primeiro match e já sai
            else:
                # Para OCR → fuzzy match                
                ratio = SequenceMatcher(None, exam_part, chave_norm).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = (chave_norm, meta)
                    best_exam_name = exam_part

        # ---------- Se encontramos exame, extrair valor ----------
        if best_match:
            chave_norm, meta = best_match
            column_title = meta["title"]

            if chave_norm in exames_processados:
                continue  # Skip if we already found this exam

            if not is_ocr:
                # Hemograma diferencial - SPECIAL HANDLING FOR BOTH FORMATS                
                if chave_norm in hemograma_abs_norm:
                    print(f"  ⚡ {chave_norm} detectado como hemograma diferencial")
                    
                    # FIRST: Try Nossa Senhora format (Andrea PDF) - THIS WORKS!
                    trecho = linha.split("(")[0].replace(" ", "")
                    m_valor = re.search(r',0([\d\.]+,\d+)', trecho)
                    
                    if m_valor:
                        # Nossa Senhora format detected
                        valor = m_valor.group(1).replace(".", "").replace(",", ".")
                        print(f"  ✅ Valor absoluto (Nossa Senhora) para {chave_norm}: {valor}")
                    else:
                        # SECOND: Try Patologia Clínica format (Jason PDF)
                        # For format: "Neutrófilos 67,2 45,0 a 70,0 2.634,2 1.500,0 a 7.000,0"
                        # We need to capture the absolute value (4th number pattern)
                        
                        # Look for the table pattern with absolute values
                        pattern_patologia = rf"{re.escape(meta['synonyms'][0])}[^\\d]*[\\d,]+\\s+[\\d,]+\\s+[\\d,]+\\s+([\\d.,]+)"
                        m_valor = re.search(pattern_patologia, linha, re.IGNORECASE)
                        
                        if m_valor:
                            # Patologia format detected
                            valor = m_valor.group(1).replace(".", "").replace(",", ".")
                            print(f"  ✅ Valor absoluto (Patologia) para {chave_norm}: {valor}")
                        else:
                            # Fallback: extract all numbers and try to identify absolute value
                            numbers = re.findall(r'(\d+[.,]?\d+)', linha)
                            if len(numbers) >= 4:
                                # In table format, absolute value is usually the 4th number
                                valor = numbers[3].replace(".", "").replace(",", ".")
                                print(f"  ✅ Valor absoluto (table position) para {chave_norm}: {valor}")

                # SPECIAL HANDLING FOR QUALITATIVE EXAMS LIKE TROPONINA
                if meta.get("qualitative", False):
                    print(f"  ⚡ Qualitative exam detected: {chave_norm}")
                    # Look for qualitative result patterns
                    for offset in range(0, 4):
                        if i + offset < len(linhas):
                            l_check = linhas[i + offset]
                            print(f"  🔍 Checking line {i+offset}: '{l_check}'")  # <-- ADD THIS
                            # Check for qualitative result patterns
                            for qual_pattern, numeric_value in qualitative_values.items():
                                if re.search(qual_pattern, l_check.upper()):
                                    valor = numeric_value
                                    print(f"  ✅ Qualitative value found: {qual_pattern} -> {valor}")
                                    break
                            if valor:
                                break
                                
                # KEEP YOUR EXISTING FALLBACK LOGIC (this is good!)
                if not valor:
                    for offset in range(0, 6):
                        if i + offset < len(linhas):
                            l = linhas[i + offset]
                            m_valor = re.search(r"Resultado[:\.]*\s*([\d]+[.,]?\d*)", l, re.IGNORECASE)
                            if m_valor:
                                valor = m_valor.group(1)
                                print(f"  ✅ Valor encontrado para {chave_norm} na linha {i+offset}: {valor}")
                                break
                    if meta.get("qualitative", False):
                        print(f"  ⚡ No numeric value found for qualitative exam {chave_norm}, trying qualitative patterns")
                        for offset in range(0, 3):
                            if i + offset < len(linhas):
                                l_check = linhas[i + offset]
                                # Look for qualitative result patterns
                                for qual_pattern, numeric_value in qualitative_values.items():
                                    if re.search(qual_pattern, l_check.upper()):  # <-- FIXED!
                                        valor = numeric_value
                                        print(f"  ✅ Qualitative value found: {qual_pattern} -> {valor}")
                                        break
                                if valor:
                                    break
                    if not valor:
                        #m_valor2 = re.search(r":\s*([\d]+[.,]?\d+)", linha)
                        m_valor2 = re.search(r'(\d+[.,]?\d+)', linha)
                        if m_valor2:
                            valor = m_valor2.group(1)
                            print(f"  ✅ Valor encontrado direto para {chave_norm}: {valor}")                    

            else:
                # OCR → talvez precise reavaliar com a linha inteira
                print(f"  best_ratio: {best_ratio}")
                if best_ratio < 0.6:
                    for chave_norm_temp, meta_temp in exames_lookup.items():
                        ratio = SequenceMatcher(None, linha_norm, chave_norm_temp).ratio()
                        if ratio > best_ratio:
                            best_ratio = ratio
                            best_match = (chave_norm_temp, meta_temp)
                            best_exam_name = "FULL_LINE"

                if best_ratio > 0.55:
                    # Procura valor na linha atual e nas próximas 5 linhas
                    for offset in range(0, 6):
                        if i + offset < len(linhas):
                            l_check = linhas[i + offset]

                            if not is_lab_result_line(l_check):
                                continue

                            all_numbers = re.findall(r"\d+[.,]?\d*", l_check)
                            print(f"  all_numbers: {all_numbers}")
                            print(f"  l_check: {l_check}")
                            #print(f"  chave_norm: {chave_norm}")
                            if all_numbers:
                                print(f"  DEBUG OCR: Números encontrados na linha {i+offset}: {all_numbers}")
                                if chave_norm in hemograma_abs_norm and len(all_numbers) >= 2:
                                    valor = all_numbers[1]
                                    print(f"  ✅ Valor absoluto (segundo número) para {chave_norm}: {valor}")
                                else:
                                    valor = all_numbers[0]
                                    print(f"  ✅ Primeiro número encontrado para {chave_norm}: {valor}")
                                break


            # ---------- Registrar resultado ----------
            if valor:
                try:
                    valor = float(str(valor).replace(",", "."))
                    print(f"  🎯 Valor final: {valor}")

                    if paciente not in dados_novos:
                        dados_novos[paciente] = []
                    dados_novos[paciente].append({
                        "Data": data,
                        "Exame": column_title,
                        "Valor": valor,
                        "Fonte": arquivo.name
                    })

                    exames_processados.add(chave_norm)  # <--- ADD THIS LINE
                    linha_usada = True  # marca a linha como já processada
                except:
                    print(f"  ❌ Erro convertendo valor: {valor}")
            else:
                print(f"  ❌ Não foi possível encontrar valor para {chave_norm}")

                        

# === CARREGAR PLANILHA EXISTENTE (SE HOUVER) ===
dados_finais = {}
if Path(saida_planilha).exists():
    xls = pd.ExcelFile(saida_planilha)
    for paciente in xls.sheet_names:
        df_old = pd.read_excel(xls, sheet_name=paciente)
        registros = []
        for _, row in df_old.iterrows():
            for exame in df_old.columns[2:]:
                if not pd.isna(row[exame]):
                    registros.append({
                        "Data": row["Data"],
                        "Exame": exame,
                        "Valor": row[exame],
                        "Fonte": row["Fonte"] if "Fonte" in df_old.columns else "Desconhecido"
                    })
        dados_finais[paciente] = registros

# === FUNDIR NOVOS DADOS ===
for paciente, registros in dados_novos.items():
    if paciente not in dados_finais:
        dados_finais[paciente] = []
    dados_finais[paciente].extend(registros)

# === REMOVER DUPLICATAS ===
for paciente in dados_finais:
    df = pd.DataFrame(dados_finais[paciente])
    df = df.drop_duplicates(subset=["Data", "Exame", "Valor", "Fonte"])
    dados_finais[paciente] = df.to_dict("records")

# === SALVAR PLANILHA ===
def sanitize_sheet_name(name: str) -> str:
    # Remove caracteres proibidos
    invalid = r'[:\\/*?[\]]'
    name = re.sub(invalid, "_", name)
    
    # If name becomes empty after sanitization, use a default
    if not name.strip():
        name = "Paciente_Desconhecido"
    
    # Limita a 31 caracteres
    return name[:31]

with pd.ExcelWriter(saida_planilha, engine="openpyxl") as writer:
    for paciente, registros in dados_finais.items():
        df = pd.DataFrame(registros)
        # Force all column names to use the title          
        df["Exame"] = df["Exame"].apply(lambda x: map_exames[x]["title"] if x in map_exames else x)
        df_wide = df.pivot_table(index=["Data", "Fonte"], columns="Exame", values="Valor", aggfunc="first").reset_index()
        df_wide = df_wide.sort_values("Data")
        
        sheet_name = sanitize_sheet_name(paciente)
        print(f"📄 Saving sheet: '{paciente}' -> '{sheet_name}'")  # Debug output
        
        df_wide.to_excel(writer, sheet_name=sheet_name, index=False)

# === INSERIR GRÁFICOS ===
wb = load_workbook(saida_planilha)
for paciente in wb.sheetnames:
    ws = wb[paciente]
    max_col = ws.max_column
    max_row = ws.max_row
    row_pos = max_row + 3

    for tema, exames in grupos.items():
        cols = []
        for col in range(3, max_col+1):
            header = ws.cell(row=1, column=col).value
            if header in exames:
                cols.append(col)
        if not cols:
            continue

        chart = LineChart()
        chart.title = tema
        chart.y_axis.title = "Valores"
        chart.x_axis.title = "Data"

        for col in cols:
            data = Reference(ws, min_col=col, max_col=col, min_row=1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)

        dates = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=max_row)
        chart.set_categories(dates)

        ws.add_chart(chart, f"A{row_pos}")
        row_pos += 15

wb.save(saida_planilha)

def main():
    # (move all your code here)
    # your final print stays the same:
    print(f"\n✅ Planilha atualizada: {saida_planilha}")

if __name__ == "__main__":
    main()
